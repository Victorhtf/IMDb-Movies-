import imdb
import openpyxl
import os

print('Iniciando script...')

try:
    ia = imdb.IMDb()
    print('Conexão com a API estabelecida.')

    # Adicionar opção de pesquisa
    opcao_pesquisa = input('Digite "1" para buscar pelos top 250 filmes ou "2" para fazer uma pesquisa livre: ')

    if opcao_pesquisa == '1':
        top250 = ia.get_top250_movies()

        # Salvar top 250 em uma planilha xlsx
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Top 250 Filmes'
        ws['A1'] = 'Título'
        ws['B1'] = 'Nota'

        for i, filme in enumerate(top250, start=2):
            titulo = filme['title']
            nota = filme['rating']
            ws[f'A{i}'] = titulo
            ws[f'B{i}'] = nota

        cwd = os.getcwd()
        wb.save(os.path.join(cwd, "top250_filmes.xlsx"))

        print(f'Top 250 filmes salvo em {os.path.join(cwd, "top250_filmes.xlsx")}')

    elif opcao_pesquisa == '2':
        pesquisa = input('Digite o nome do filme que deseja pesquisar: ')
        resultados = ia.search_movie(pesquisa)

        # Selecionar o primeiro resultado da pesquisa
        id_filme = resultados[0].getID()
        filme = ia.get_movie(id_filme)
        titulo = filme['title']
        nota = filme['rating']
        print(f'Filme encontrado: {titulo} (nota: {nota})')

        # Salvar resultado em uma planilha xlsx
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Filmes'
        ws['A1'] = 'Título'
        ws['B1'] = 'Nota'
        ws['A2'] = titulo
        ws['B2'] = nota

        cwd = os.getcwd()
        wb.save(os.path.join(cwd, "filmes.xlsx"))

        print(f'Filme salvo em {os.path.join(cwd, "filmes.xlsx")}')

    else:
        print('Opção inválida. Tente novamente.')

except imdb.IMDbError as e:
    print(f'Erro: {e}')

print('Script finalizado.')
